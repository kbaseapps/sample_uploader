
import os
import unittest
import json
from configparser import ConfigParser
from openpyxl import load_workbook
import copy
import shutil
from sample_uploader.authclient import KBaseAuth as _KBaseAuth

from sample_uploader.utils.importer import import_samples_from_file, find_header_row
from sample_uploader.utils.mappings import SESAR_mappings, ENIGMA_mappings, aliases
from sample_uploader.utils.sample_utils import get_sample


class sample_uploaderTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.token = os.environ.get('KB_AUTH_TOKEN', None)
        config_file = os.environ.get('KB_DEPLOYMENT_CONFIG', None)
        cls.cfg = {}
        config = ConfigParser()
        config.read(config_file)
        for nameval in config.items('sample_uploader'):
            cls.cfg[nameval[0]] = nameval[1]
        authServiceUrl = cls.cfg['auth-service-url']
        auth_client = _KBaseAuth(authServiceUrl)
        cls.username = auth_client.get_user(cls.token)
        cls.test_dir = os.path.dirname(os.path.dirname(os.path.realpath(__file__)))
        sample_server = cls.cfg.get('sample-server', 'sampleservice')
        cls.sample_url = cls.cfg['kbase-endpoint'] + '/{}'.format(sample_server)
        cls.workspace_url = cls.cfg['workspace-url']
        cls.callback_url = os.environ['SDK_CALLBACK_URL']

    def _verify_samples(self, sample_set, compare_path):
        with open(compare_path) as f:
            compare = json.load(f)

        for it, samp in enumerate(sample_set['samples']):
            sample = get_sample(samp, self.sample_url, self.token)
            self._compare_sample(sample, compare[it])

    def _compare_sample(self, s, sc, check_version=True, check_id=False):
        self.assertEqual(s['name'], sc['name'], msg=f"s: {json.dumps(s['name'])}\nsc: {json.dumps(sc['name'])}")
        # self.assertEqual(s['node_tree'], sc['node_tree'], msg=f"s: {json.dumps(s['node_tree'])}\nsc: {json.dumps(sc['node_tree'])}")
        for i, s_node in enumerate(s['node_tree']):
            sc_node = sc['node_tree'][i]
            # check all fields except 'source_meta'
            if 'source_meta' in s_node:
                s_node.pop('source_meta')
            if 'source_meta' in sc_node:
                sc_node.pop('source_meta')
            self.assertEqual(s_node, sc_node, msg=f"s: {json.dumps(s_node)}\nsc: {json.dumps(sc_node)}")
        if check_version:
            self.assertEqual(s['version'], sc['version'], msg=f"s: {json.dumps(s['version'])}\nsc: {json.dumps(sc['version'])}")
        if check_id:
            self.assertEqual(s['id'], sc['id'])

    def test_find_header_row(self):
        # test enigma files
        sample_file = os.path.join(self.test_dir, 'data', 'fake_samples_ENIGMA.xlsx')
        header_row = find_header_row(sample_file, 'ENIGMA')
        self.assertEqual(header_row, 0)

        # test sesar tsv files
        sample_file = os.path.join(self.test_dir, 'data', 'fake_samples.tsv')
        header_row = find_header_row(sample_file, 'SESAR')
        self.assertEqual(header_row, 1)

        sample_file = os.path.join(self.test_dir, 'example_data', 'samples_all.tsv')
        header_row = find_header_row(sample_file, 'SESAR')
        self.assertEqual(header_row, 0)

        # test sesar xlsx files
        sample_file = os.path.join(self.test_dir, 'example_data', 'ANLPW_JulySamples_IGSN_v2.xls')
        header_row = find_header_row(sample_file, 'SESAR')
        self.assertEqual(header_row, 1)

        sample_file = os.path.join(self.test_dir, 'example_data', 'SESAR_Torn_biochem_SFA.xlsx')
        header_row = find_header_row(sample_file, 'SESAR')
        self.assertEqual(header_row, 0)

        # test sesar csv files
        sample_file = os.path.join(self.test_dir, 'example_data', 'isgn_sample_example.csv')
        header_row = find_header_row(sample_file, 'SESAR')
        self.assertEqual(header_row, 1)

        sample_file = os.path.join(self.test_dir, 'example_data', 'ncbi_sample_example.csv')
        header_row = find_header_row(sample_file, 'SESAR')
        self.assertEqual(header_row, 0)

    # @unittest.skip('x')
    def test_ENIGMA_format(self):
        # test default sample server
        ori_sample_file = os.path.join(self.test_dir, 'data', 'fake_samples_ENIGMA.xlsx')
        sample_file = os.path.join(self.test_dir, 'data', 'updated_fake_samples_ENIGMA.xlsx')
        shutil.copy2(ori_sample_file, sample_file)

        params = {
            'workspace_name': 'workspace_name',
            'sample_file': sample_file,
            'file_format': "ENIGMA",
            'prevalidate': 1,
        }

        header_row_index = 1

        mappings = {'enigma': ENIGMA_mappings, 'sesar': SESAR_mappings, 'kbase': {}}
        sample_set, errors, sample_data_json = import_samples_from_file(
            params,
            self.sample_url,
            self.workspace_url,
            self.callback_url,
            self.username,
            self.token,
            mappings[str(params.get('file_format')).lower()].get('groups', []),
            mappings[str(params.get('file_format')).lower()].get('date_columns', []),
            mappings[str(params.get('file_format')).lower()].get('column_unit_regex', []),
            {"samples": []},
            header_row_index,
            aliases.get(params.get('file_format').lower(), {})
        )

        samples = sample_set['samples']
        self.assertEqual(len(samples), 3)
        expected_sample_name = ['s1', 's2', 's3']
        self.assertCountEqual([sample['name'] for sample in samples], expected_sample_name)
        self.assertEqual(len(errors), 0)

        compare_path = os.path.join(self.test_dir, "data", "fake_samples_ENIGMA.json")
        self._verify_samples(sample_set, compare_path)

        # test updating samples

        # test updating samples with same sample file
        expected_error = "No sample is produced from file.\nThe input sample set has identical information to the input file"
        with self.assertRaisesRegex(ValueError, expected_error):
            import_samples_from_file(
                params,
                self.sample_url,
                self.workspace_url,
                self.callback_url,
                self.username,
                self.token,
                mappings[str(params.get('file_format')).lower()].get('groups', []),
                mappings[str(params.get('file_format')).lower()].get('date_columns', []),
                mappings[str(params.get('file_format')).lower()].get('column_unit_regex', []),
                sample_set,
                header_row_index,
                aliases.get(params.get('file_format').lower(), {})
            )

        # test updating a single value
        wb = load_workbook(sample_file)
        ws = wb.active
        new_latitude = 66.6
        self.assertNotEqual(new_latitude, ws['E3'].value)
        ws['E3'].value = new_latitude  # update latitude value to 66.6 for S1
        wb.save(sample_file)

        sample_set, errors, sample_data_json = import_samples_from_file(
            params,
            self.sample_url,
            self.workspace_url,
            self.callback_url,
            self.username,
            self.token,
            mappings[str(params.get('file_format')).lower()].get('groups', []),
            mappings[str(params.get('file_format')).lower()].get('date_columns', []),
            mappings[str(params.get('file_format')).lower()].get('column_unit_regex', []),
            sample_set,
            header_row_index,
            aliases.get(params.get('file_format').lower(), {})
        )

        updated_samples = sample_set['samples']
        self.assertEqual(len(updated_samples), 3)
        expected_sample_name = ['s1', 's2', 's3']
        self.assertCountEqual([sample['name'] for sample in updated_samples], expected_sample_name)
        self.assertEqual(len(errors), 0)

        with open(compare_path) as f:
            data = json.load(f)

        data[0]['version'] += 1  # sample version should be bumped
        data[0]['node_tree'][0]['meta_controlled']['latitude']['value'] = new_latitude

        with open(compare_path, 'w') as json_file:
            json.dump(data, json_file)

        self._verify_samples(sample_set, compare_path)

        # test adding a column
        wb = load_workbook(sample_file)
        ws = wb.active
        new_column = '?size?'  # add a user metadata column
        ws['I2'].value = new_column
        size_load = 10
        for cell in ws['I3':'I5']:
            cell[0].value = size_load
        wb.save(sample_file)

        sample_set, errors, sample_data_json = import_samples_from_file(
            params,
            self.sample_url,
            self.workspace_url,
            self.callback_url,
            self.username,
            self.token,
            mappings[str(params.get('file_format')).lower()].get('groups', []),
            mappings[str(params.get('file_format')).lower()].get('date_columns', []),
            mappings[str(params.get('file_format')).lower()].get('column_unit_regex', []),
            sample_set,
            header_row_index,
            aliases.get(params.get('file_format').lower(), {})
        )

        updated_samples = sample_set['samples']
        self.assertEqual(len(updated_samples), 3)
        expected_sample_name = ['s1', 's2', 's3']
        self.assertCountEqual([sample['name'] for sample in updated_samples], expected_sample_name)
        self.assertEqual(len(errors), 0)

        with open(compare_path) as f:
            data = json.load(f)

        for sample in data:
            sample['version'] += 1  # sample version should be bumped
            sample['node_tree'][0]['meta_user'][new_column] = {'value': size_load}  # a new user meta data should be added

        with open(compare_path, 'w') as json_file:
            json.dump(data, json_file)
        self._verify_samples(sample_set, compare_path)

        # test adding a new sample (row)
        wb = load_workbook(sample_file)
        ws = wb.active
        for cell in ws[5]:
            ws[cell.column_letter + '6'] = cell.value  # copy s3 (line 5) to the next line
        new_sample = 's4'
        ws['A6'].value = new_sample  # update the sample id for the new row
        wb.save(sample_file)

        sample_set, errors, sample_data_json = import_samples_from_file(
            params,
            self.sample_url,
            self.workspace_url,
            self.callback_url,
            self.username,
            self.token,
            mappings[str(params.get('file_format')).lower()].get('groups', []),
            mappings[str(params.get('file_format')).lower()].get('date_columns', []),
            mappings[str(params.get('file_format')).lower()].get('column_unit_regex', []),
            sample_set,
            header_row_index,
            aliases.get(params.get('file_format').lower(), {})
        )

        updated_samples = sample_set['samples']
        self.assertEqual(len(updated_samples), 4)
        expected_sample_name = ['s1', 's2', 's3', new_sample]
        self.assertCountEqual([sample['name'] for sample in updated_samples], expected_sample_name)
        self.assertEqual(len(errors), 0)

        with open(compare_path) as f:
            data = json.load(f)

        data.insert(0, copy.deepcopy(data[-1]))  # copy S3 data and update sample ID and version
        data[0]['name'] = new_sample
        data[0]['node_tree'][0]['id'] = new_sample
        data[0]['version'] = 1

        with open(compare_path, 'w') as json_file:
            json.dump(data, json_file)

        self._verify_samples(sample_set, compare_path)

    # @unittest.skip('x')
    def test_import_SESAR_format(self):
        # test default sample server
        sample_file = os.path.join(self.test_dir, 'data', 'fake_samples.tsv')

        params = {
            'workspace_name': 'workspace_name',
            'sample_file': sample_file,
            'file_format': "sesar",
            'name_field': 'test name field',
            'prevalidate': 1
        }
        header_row_index = 1

        mappings = {'enigma': ENIGMA_mappings, 'sesar': SESAR_mappings, 'kbase': {}}
        sample_set, errors, sample_data_json = import_samples_from_file(
            params,
            self.sample_url,
            self.workspace_url,
            self.callback_url,
            self.username,
            self.token,
            mappings[str(params.get('file_format')).lower()].get('groups', []),
            mappings[str(params.get('file_format')).lower()].get('date_columns', []),
            mappings[str(params.get('file_format')).lower()].get('column_unit_regex', []),
            {"samples": []},
            header_row_index,
            aliases.get(params.get('file_format').lower(), {})
        )

        samples = sample_set['samples']
        self.assertEqual(len(samples), 3)
        expected_sample_name = ['s1', 's2', 's3']
        self.assertCountEqual([sample['name'] for sample in samples], expected_sample_name)
        self.assertEqual(len(errors), 0)

        compare_path = os.path.join(self.test_dir, "data", "fake_samples.json")
        self._verify_samples(sample_set, compare_path)

    # @unittest.skip('x')
    def test_KBASE_format(self):
        # test default sample server
        sample_file = os.path.join(self.test_dir, 'example_data', 'ncbi_sample_example.csv')
        params = {
            'workspace_name': 'workspace_name',
            'sample_file': sample_file,
            'file_format': "KBASE",
            'id_field': 'id',
            'prevalidate': 1,
        }
        header_row_index = 0

        mappings = {'enigma': ENIGMA_mappings, 'sesar': SESAR_mappings, 'kbase': {}}
        sample_set, errors, sample_data_json = import_samples_from_file(
            params,
            self.sample_url,
            self.workspace_url,
            self.callback_url,
            self.username,
            self.token,
            mappings[str(params.get('file_format')).lower()].get('groups', []),
            mappings[str(params.get('file_format')).lower()].get('date_columns', []),
            mappings[str(params.get('file_format')).lower()].get('column_unit_regex', []),
            {"samples": []},
            header_row_index,
            aliases.get(params.get('file_format').lower(), {})
        )

        samples = sample_set['samples']
        self.assertEqual(len(samples), 2)
        expected_sample_name = ['SAMN03166112', 'SAMN04383980']
        self.assertCountEqual([sample['name'] for sample in samples], expected_sample_name)
        self.assertEqual(len(errors), 0)
